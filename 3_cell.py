from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# ws["A1"] = 1
# ws["B3"] = 4

# print(ws.cell(row=1, column=1)) #=(1.1)
# wb.save("sample3.xlsx")

from random import *

index=1
for x in range(1,11):
		for y in range(1,11):
				ws.cell(row=x, column=y, value=index)
				index += 1
# ws.cell(row=x, column=y, value=randint(0,100)) # random 100 datas 

wb.save("sample3.xlsx")