from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # make new sheet by formal
ws.title = "Mysheet"
ws.sheet_properties.tabColor = "44dd1d"

ws1 = wb.create_sheet("YourSheet") # make sheet given name 
ws2 = wb.create_sheet("NewSheet", 2) #index setting

new_ws = wb["NewSheet"] # make dictionary can access to Sheet 

print(wb.sheetnames) # every sheet name 

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied"
# Copy sheet 


wb.save("sample.xlsx")
