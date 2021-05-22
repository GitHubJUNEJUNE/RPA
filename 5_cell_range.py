from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["num", "eng", "math"])
for i in range(1,11):
    ws.append([i, randint(1,100), randint(1,100)])

# col_eng = ws["B"] # eng column
# print(col_eng)
# for cell in col_eng:
#     print(cell.value)

# col_range = ws["B:C"] # BC 열 데이터 

# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_range = ws[2:6] #2~6번째 줄 가져오기 
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()
#[1]1행 [A]1열


from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
        print(cell.coordinate, end=" ")#cell's 좌표정보 가져올수 있음 
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ") # 튜플 형식으로 각 좌표 ('x',y)로 나타남 
        print(xy[0], end="") #=A,B,C
        print(xy[1], end=" ") # =1,2,3
    print()


wb.save("sample5.xlsx")

